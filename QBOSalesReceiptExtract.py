'''
QBOSalesReceiptExtract by EccentricWorkshop (ecc.ws or github.com/EccentricWkshp)
Initial on 06/05/2024

Connects to QBO API.
Requires QBO developer account, creation of an app, and generation of OAuth credentials.
Gets sales receipts for specified number of days.
Extracts date, name, state or country, total amount, shipping amount, and SKUs along with quantity if greater than 1.
Saves everything to sales_receipts.xlsx in the current directory.
Has debug feature for BillAddr, ShipAddr, and sales_receipts.

QBO API Docs: https://developer.intuit.com/app/developer/qbo/docs/api/accounting/all-entities/salesreceipt
'''

import json
import argparse
from datetime import datetime, timedelta
import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pycountry

# Load configuration
with open('config.json', 'r') as config_file:
    config = json.load(config_file)

# Argument parser setup
parser = argparse.ArgumentParser(description='Fetch sales receipts from QuickBooks Online.')
parser.add_argument('--days', type=int, default=config.get('default_days', 30), help='Number of days of sales receipts to parse. Default is 30 days.')
args = parser.parse_args()

# Calculate the date range
end_date = datetime.now().date()
start_date = end_date - timedelta(days=args.days)

# Function to refresh access token
def refresh_access_token():
    token_url = 'https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer'
    auth = HTTPBasicAuth(config['client_id'], config['client_secret'])
    response = requests.post(token_url, auth=auth, data={
        'grant_type': 'refresh_token',
        'refresh_token': config['refresh_token']
    })
    if response.status_code == 200:
        token_info = response.json()
        return token_info['access_token']
    else:
        print("Error refreshing token:", response.json())
        return None

# Set up OAuth2 session
access_token = refresh_access_token()
if not access_token:
    exit()

# Define the headers
headers = {
    'Authorization': f'Bearer {access_token}',
    'Accept': 'application/json',
    'Content-Type': 'application/json'
}

# Fetch sales receipts
def fetch_sales_receipts():
    query = f"select * from SalesReceipt where TxnDate >= '{start_date}' and TxnDate <= '{end_date}'"
    url = f"https://{('sandbox-' if config['sandbox'] else '')}quickbooks.api.intuit.com/v3/company/{config['realm_id']}/query"
    try:
        response = requests.get(url, headers=headers, params={'query': query})
        response.raise_for_status()
        return response.json().get('QueryResponse', {}).get('SalesReceipt', [])
    except requests.exceptions.RequestException as e:
        print("Error fetching sales receipts:", e)
        return []

sales_receipts = fetch_sales_receipts()

# Write sales receipts JSON to file if receipt_debug is enabled
if config.get('receipt_debug', False):
    with open('sales_receipts.json', 'w') as f:
        json.dump(sales_receipts, f, indent=4)

# Generate a list of country names and abbreviations from pycountry
country_names = {country.name.lower(): country.alpha_2 for country in pycountry.countries}
country_abbreviations = {country.alpha_2: country.name for country in pycountry.countries}

# Generate a list of US states and abbreviations
us_states = {state.name: state.code.split('-')[-1] for state in pycountry.subdivisions.get(country_code='US')}
us_state_abbreviations = {state.code.split('-')[-1]: state.name for state in pycountry.subdivisions.get(country_code='US')}

# Extract state or country from address lines
def extract_state_or_country(address):
    state = 'Unknown'
    country = 'Unknown'
    city_state_postal = None
    for line in ['Line3', 'Line4', 'Line5']:
        if line in address and address[line]:
            parts = address[line].split()
            if len(parts) >= 2:
                # Check for state abbreviations and names
                if parts[-2] in us_state_abbreviations:
                    state = parts[-2]
                    if state == 'WA':
                        city_state_postal = address[line]
                elif ' '.join(parts[-2:]) in us_states:
                    state = us_states[' '.join(parts[-2:])]
                # Check for country abbreviations and names
                elif parts[-1] in country_abbreviations:
                    country = country_abbreviations[parts[-1]]
                elif ' '.join(parts).lower() in country_names:
                    country = country_abbreviations[country_names[' '.join(parts).lower()]]
            elif len(parts) >= 1:
                possible_country = ' '.join(parts).lower()
                if possible_country in country_names:
                    country = country_abbreviations[country_names[possible_country]]
    if state == 'WA' and city_state_postal:
        return city_state_postal
    return state if state != 'Unknown' else country

# Extract receipt details
data = {}
for receipt in sales_receipts:
    shipping_cost = 0
    total_amount = receipt['TotalAmt']
    txn_date = receipt['TxnDate']
    customer = receipt['CustomerRef']['name']
    ship_addr = receipt.get('ShipAddr', {})
    bill_addr = receipt.get('BillAddr', {})

    state_or_country = extract_state_or_country(ship_addr)

    # Debug printout for addresses
    if config.get('address_debug', False):
        print(f"Billing Address: {bill_addr}")
        print(f"Shipping Address: {ship_addr}")

    receipt_key = (txn_date, customer, state_or_country)

    if receipt_key not in data:
        data[receipt_key] = {
            "Date": txn_date,
            "Customer": customer,
            "State": state_or_country,
            "Total Amount": total_amount,
            "Shipping Cost": shipping_cost,
            "SKUs": []
        }

    for line in receipt.get('Line', []):
        if line.get('DetailType') == "SalesItemLineDetail":
            item_ref = line['SalesItemLineDetail']['ItemRef']
            sku_name = item_ref.get('name', '').split(':')[-1].strip()  # Get the name portion after the colon
            quantity = line['SalesItemLineDetail'].get('Qty', 1)
            if sku_name and sku_name != 'Unknown':  # Skip unknown SKUs
                if quantity > 1:
                    sku_name = f"{quantity}x {sku_name}"
                data[receipt_key]["SKUs"].append(sku_name)
            # Check for shipping item and update shipping_cost
            if item_ref['value'] == 'SHIPPING_ITEM_ID':
                shipping_cost += line['Amount']

    data[receipt_key]['Shipping Cost'] = shipping_cost

# Convert to DataFrame
df_data = []
for key, value in data.items():
    value["SKUs"] = '; '.join(value["SKUs"])
    df_data.append(value)

df = pd.DataFrame(df_data)

# Save to XLSX
file_path = "sales_receipts.xlsx"
df.to_excel(file_path, index=False)

# Adjust column widths
wb = load_workbook(file_path)
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

wb.save(file_path)

print(f"Sales receipts from the last {args.days} days have been saved to {file_path}")
